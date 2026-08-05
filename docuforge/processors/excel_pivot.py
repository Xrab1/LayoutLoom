"""Portable pivot-style Excel summaries with a native Excel upgrade path."""

from __future__ import annotations

import math
from collections import defaultdict
from pathlib import Path
from typing import Any, Sequence

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple

from ..models import MissingEngineError, ValidationError
from ..utils import atomic_output, ensure_output_dir, safe_filename, unique_path

PathLike = str | Path
_FUNCTIONS = {"sum", "count", "average", "max", "min"}


def _normalize_fields(values: Sequence[str], label: str) -> tuple[str, ...]:
    fields = tuple(str(value).strip() for value in values if str(value).strip())
    folded = [value.casefold() for value in fields]
    if len(folded) != len(set(folded)):
        raise ValidationError(f"{label} 不能包含重复字段")
    return fields


def _aggregate(values: list[Any], function: str, row_count: int) -> float | int:
    present = [value for value in values if value not in (None, "")]
    if function == "count":
        return len(present) if present else row_count
    numbers: list[float] = []
    for value in present:
        if isinstance(value, bool):
            raise ValidationError("数值字段包含布尔值，无法执行数值汇总")
        try:
            number = float(value)
        except (TypeError, ValueError, OverflowError) as exc:
            raise ValidationError(f"数值字段包含非数字内容：{value}") from exc
        if not math.isfinite(number):
            raise ValidationError("数值字段包含无穷大或非数字值")
        numbers.append(number)
    if not numbers:
        return 0
    if function == "sum":
        result = sum(numbers)
    elif function == "average":
        result = sum(numbers) / len(numbers)
    elif function == "max":
        result = max(numbers)
    else:
        result = min(numbers)
    return int(result) if result.is_integer() else result


def _column_label(fields: Sequence[str], key: tuple[Any, ...]) -> str:
    return " / ".join(
        f"{field}={'' if value is None else value}"
        for field, value in zip(fields, key)
    ) or "汇总"


def excel_create_pivot_summary(
    source: PathLike,
    output_dir: PathLike,
    source_sheet: str,
    source_range: str,
    target_sheet: str = "数据透视表",
    target_cell: str = "A1",
    row_fields: Sequence[str] = (),
    column_fields: Sequence[str] = (),
    data_field: str | None = None,
    function: str = "sum",
    overwrite: bool = False,
) -> list[Path]:
    """Create a static summary that opens consistently in Excel, WPS and LO."""

    from .office import _load_workbook

    source_path, workbook = _load_workbook(source)
    source_sheet_name = str(source_sheet).strip()
    target_sheet_name = str(target_sheet).strip()
    if source_sheet_name not in workbook.sheetnames:
        raise ValidationError(f"source_sheet 不存在：{source_sheet_name}")
    if (
        not target_sheet_name
        or len(target_sheet_name) > 31
        or any(character in target_sheet_name for character in "[]:*?/\\")
    ):
        raise ValidationError("target_sheet 不是有效的 Excel 工作表名")
    if source_sheet_name.casefold() == target_sheet_name.casefold():
        raise ValidationError("兼容汇总模式要求目标工作表与数据源工作表不同")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(str(source_range))
    except (TypeError, ValueError) as exc:
        raise ValidationError(f"source_range 不是有效的 A1 区域：{source_range}") from exc
    if min_row >= max_row:
        raise ValidationError("source_range 至少需要一行标题和一行数据")
    try:
        target_row, target_col = coordinate_to_tuple(str(target_cell).strip().upper())
    except (TypeError, ValueError) as exc:
        raise ValidationError(f"target_cell 不是有效单元格：{target_cell}") from exc

    worksheet = workbook[source_sheet_name]
    headers = [
        str(worksheet.cell(min_row, column).value or "").strip()
        for column in range(min_col, max_col + 1)
    ]
    if any(not header for header in headers):
        raise ValidationError("source_range 的标题行不能包含空白字段名")
    folded_headers = [header.casefold() for header in headers]
    if len(folded_headers) != len(set(folded_headers)):
        raise ValidationError("source_range 的标题行不能包含重复字段名")
    header_index = {header.casefold(): index for index, header in enumerate(headers)}

    rows = _normalize_fields(row_fields, "row_fields")
    columns = _normalize_fields(column_fields, "column_fields")
    if {value.casefold() for value in rows} & {
        value.casefold() for value in columns
    }:
        raise ValidationError("同一字段不能同时作为行字段和列字段")
    data_name = str(data_field).strip() if data_field is not None else ""
    requested = (*rows, *columns, *([data_name] if data_name else []))
    missing = [value for value in requested if value.casefold() not in header_index]
    if missing:
        raise ValidationError(f"字段不存在：{missing[0]}")
    if not rows and not columns and not data_name:
        raise ValidationError("至少需要一个行字段、列字段或数据字段")
    function_name = str(function).strip().lower()
    if function_name not in _FUNCTIONS:
        raise ValidationError("function 必须是 sum、count、average、max 或 min")
    if not data_name:
        function_name = "count"

    source_rows = [
        tuple(
            worksheet.cell(row_number, column).value
            for column in range(min_col, max_col + 1)
        )
        for row_number in range(min_row + 1, max_row + 1)
    ]
    source_rows = [
        values for values in source_rows if any(value not in (None, "") for value in values)
    ]
    if not source_rows:
        raise ValidationError("source_range 中没有可汇总的数据行")

    row_indexes = tuple(header_index[field.casefold()] for field in rows)
    column_indexes = tuple(header_index[field.casefold()] for field in columns)
    data_index = header_index[data_name.casefold()] if data_name else None
    grouped_values: dict[tuple[tuple[Any, ...], tuple[Any, ...]], list[Any]] = defaultdict(list)
    grouped_counts: dict[tuple[tuple[Any, ...], tuple[Any, ...]], int] = defaultdict(int)
    for values in source_rows:
        row_key = tuple(values[index] for index in row_indexes)
        column_key = tuple(values[index] for index in column_indexes)
        key = (row_key, column_key)
        grouped_counts[key] += 1
        grouped_values[key].append(values[data_index] if data_index is not None else 1)

    # Preserve first-seen order instead of applying an artificial text sort.
    # This matches the source table more closely and avoids surprising Chinese,
    # numeric and date label reordering across Excel/WPS/LibreOffice.
    row_keys = list(dict.fromkeys(key[0] for key in grouped_values))
    column_keys = list(dict.fromkeys(key[1] for key in grouped_values))
    if target_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[target_sheet_name])
    target = workbook.create_sheet(target_sheet_name)

    headers_out = list(rows) if rows else ["分组"]
    if columns:
        headers_out.extend(_column_label(columns, key) for key in column_keys)
        headers_out.append("总计")
    else:
        headers_out.append(f"{function_name}_{data_name}" if data_name else "记录数")
    for offset, value in enumerate(headers_out):
        cell = target.cell(target_row, target_col + offset, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F6BED")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    output_row = target_row + 1
    for row_key in row_keys:
        displayed_row_key = row_key if rows else ("全部",)
        for offset, value in enumerate(displayed_row_key):
            target.cell(output_row, target_col + offset, value)
        value_col = target_col + len(displayed_row_key)
        if columns:
            total_values: list[Any] = []
            total_count = 0
            for offset, column_key in enumerate(column_keys):
                key = (row_key, column_key)
                count = grouped_counts.get(key, 0)
                values = grouped_values.get(key, [])
                target.cell(
                    output_row,
                    value_col + offset,
                    _aggregate(values, function_name, count) if count else 0,
                )
                total_values.extend(values)
                total_count += count
            target.cell(
                output_row,
                value_col + len(column_keys),
                _aggregate(total_values, function_name, total_count),
            )
        else:
            key = (row_key, ())
            target.cell(
                output_row,
                value_col,
                _aggregate(grouped_values[key], function_name, grouped_counts[key]),
            )
        output_row += 1

    target.freeze_panes = target.cell(target_row + 1, target_col + 1)
    for column in range(target_col, target_col + len(headers_out)):
        contents = [target.cell(row, column).value for row in range(target_row, output_row)]
        target.column_dimensions[get_column_letter(column)].width = min(
            42, max(10, max(len(str(value or "")) for value in contents) + 2)
        )

    directory = ensure_output_dir(output_dir)
    output_path = unique_path(
        directory
        / f"{source_path.stem}_{safe_filename('数据透视汇总')}{source_path.suffix}",
        overwrite,
    )
    with atomic_output(output_path) as temporary:
        workbook.save(temporary)
    return [output_path]


def excel_create_pivot_compatible(*args: Any, **kwargs: Any) -> list[Path]:
    """Use native Excel when available, otherwise create a static summary."""

    try:
        from .office import detect_office_engines

        if detect_office_engines()["microsoft_excel"].available:
            from .office_com import excel_create_pivot

            try:
                return excel_create_pivot(*args, **kwargs)
            except MissingEngineError:
                pass
    except (ImportError, KeyError, OSError):
        pass
    return excel_create_pivot_summary(*args, **kwargs)


__all__ = ["excel_create_pivot_compatible", "excel_create_pivot_summary"]

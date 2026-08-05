from pathlib import Path

from openpyxl import Workbook, load_workbook

from docuforge.processors.excel_pivot import excel_create_pivot_summary


def test_portable_pivot_summary_groups_rows_and_columns(tmp_path: Path) -> None:
    source = tmp_path / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(("地区", "季度", "金额"))
    sheet.append(("华东", "一季度", 10))
    sheet.append(("华东", "二季度", 20))
    sheet.append(("华南", "一季度", 5))
    sheet.append(("华南", "二季度", 15))
    workbook.save(source)

    output = excel_create_pivot_summary(
        source,
        tmp_path / "out",
        source_sheet="Sheet1",
        source_range="A1:C5",
        row_fields=("地区",),
        column_fields=("季度",),
        data_field="金额",
        function="sum",
    )[0]

    result = load_workbook(output, data_only=True)["数据透视表"]
    assert [result.cell(1, column).value for column in range(1, 5)] == [
        "地区",
        "季度=一季度",
        "季度=二季度",
        "总计",
    ]
    assert [result.cell(2, column).value for column in range(1, 5)] == [
        "华东",
        10,
        20,
        30,
    ]
    assert [result.cell(3, column).value for column in range(1, 5)] == [
        "华南",
        5,
        15,
        20,
    ]


def test_portable_pivot_without_data_field_counts_records(tmp_path: Path) -> None:
    source = tmp_path / "records.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("分类", "内容"))
    sheet.append(("A", "一"))
    sheet.append(("A", "二"))
    sheet.append(("B", "三"))
    workbook.save(source)

    output = excel_create_pivot_summary(
        source,
        tmp_path / "out",
        source_sheet="Sheet",
        source_range="A1:B4",
        row_fields=("分类",),
    )[0]

    result = load_workbook(output, data_only=True)["数据透视表"]
    assert result["A2"].value == "A"
    assert result["B2"].value == 2
    assert result["A3"].value == "B"
    assert result["B3"].value == 1
